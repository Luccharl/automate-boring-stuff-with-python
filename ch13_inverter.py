import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('blank.xlsx')
sheet = wb.active
sheet2 = wb.create_sheet()

produce = []

for y in range(1, sheet.max_column+1):
    for x in range(1, sheet.max_row+1):
        val = sheet.cell(row=x, column= y).value
        letter = get_column_letter(x)
        sheet2[letter + str(y)] = val
        
wb.save('blank.xlsx')