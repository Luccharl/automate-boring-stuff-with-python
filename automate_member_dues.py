import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}

#GET THE EMAIL OF THE UNPAID MEMBERS
for data in range(2, sheet.max_row+1):
    payment = sheet.cell(row=data, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=data, column=1).value
        email = sheet.cell(row=data, column=2).value
        unpaid_members[name] = email

#CONNECT AND LOGIN TO EMAIL ACCOUNT
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('lucdato@my.cspc.edu.ph', '')#sys.argv[1]

for name,email in unpaid_members.items():
    body = f'''Subject: {latest_month} dues unpaid. 
    
            Dear {name},
                Records show that you have not paid dues for {latest_month}.
                Please make this payment as soon as possible. Thank you!'''
    print(f'Sending email to {email}')
    
    send_mail_status = smtp_obj.sendmail('lucdato@my.cspc.edu.ph', email, body)
    
    if send_mail_status != {}:
        print(f'There was a problem sending email to {email}: {send_mail_status}')

smtp_obj.quit()