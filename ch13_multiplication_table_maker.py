from operator import le
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

n = 6

for i in range(1,n+1):
    sheet['A' + str(i+1)] = i
    letter = get_column_letter(i+1)
    sheet[letter + str(1)] = i
    for x in range(1,n+1):
        letter = get_column_letter(x+1)
        sheet[letter + str(i+1)] = f'= PRODUCT(A{i+1},{letter}1)'
        
wb.save('multiplication_copy.xlsx')
    