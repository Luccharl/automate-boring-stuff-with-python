import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb.active

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for rows in range(2, sheet.max_row):
    product = sheet.cell(row=rows, column=1).value
    if product in PRICE_UPDATES:
        sheet.cell(row=rows, column=2).value = PRICE_UPDATES[product]
        
wb.save('updated_produce_sales.xlsx')