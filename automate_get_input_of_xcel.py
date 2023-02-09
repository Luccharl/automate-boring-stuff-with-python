import openpyxl

def workbook(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    
    students = []
    
    
    for r in range(2, sheet.max_row+1):
        data = {}
        
        name = sheet.cell(row=r, column=1).value
        section = sheet.cell(row=r, column=2).value
        year = sheet.cell(row=r, column=3).value
        
        data['name'] = name
        data['section'] = section
        data['year'] = year
            
        students.append(data)
                
    return students
            
        

print(workbook('xcel_names.xlsx'))