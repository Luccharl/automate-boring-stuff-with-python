import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('updated_produce_sales.xlsx')
wb2 = openpyxl.Workbook()
wbsheet = wb.active
wb2sheet = wb2.active

n = 3
m = 2

for n_row in range(1,n+1):
    for line in range(1, wbsheet.max_column):
        value = wbsheet.cell(row=n_row, column=line).value
        letter = get_column_letter(line)
        wb2sheet[letter + str(n_row)] = value
        
for n_row in range(wb2sheet.max_row+1,11):
    for line in range(1, wbsheet.max_column):
        value = wbsheet.cell(row=n_row, column=line).value
        letter = get_column_letter(line)
        wb2sheet[letter + str(n_row+m)] = value

wb2.save('blank.xlsx')



